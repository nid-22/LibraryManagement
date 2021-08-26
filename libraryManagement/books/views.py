from .models import Books
from django.core.exceptions import ObjectDoesNotExist
from .forms import addBookForm,addAuthorForm
from django.shortcuts import redirect, render
import openpyxl
from .models import Books,Author
from django.contrib import messages
from django.contrib.auth.decorators import login_required
# Create your views here.
@login_required
def addExcel(request):
    if request.method =="POST":      
        excel_file = request.FILES["excel_file"]

        # you may put validations here to check extension or file size

        wb = openpyxl.load_workbook(excel_file)
        sheets = wb.sheetnames
        print(sheets)
        # getting a particular sheet by name out of many sheets
        worksheet = wb["books"]
        print(worksheet)

        excel_data = list()
        # iterating over the rows and
        # getting value from each cell in row
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
            excel_data.append(row_data)
        if excel_data[0]!=['Title', 'Description', 'AuthorID', 'Rating']:
            messages.error(request,'Please upload excel with column heads- Title,Description,AuthorID,Rating')       
            return render(request, 'Books/addExcel.html',)
        for list_ in excel_data[1:]:
            try:
                b = Books(Title = list_[0],Description =list_[1],Rating = list_[3])
                a = Author.objects.get(id = int(list_[2]) )
                b.Author = a
                b.save()
            except Exception as e:
                print(e)
        messages.success(request,'Excel data uploaded to books database')       
        return render(request, 'Books/addExcel.html',)
    return render(request, 'Books/addExcel.html')

@login_required
def addBooks(request):
    if request.method =='POST':
        bookform = addBookForm(request.POST, request.FILES) 
        authorform = addAuthorForm(request.POST)
        if bookform.is_valid(): 
            book = bookform.save()
            book.save()
            return redirect('addBooks')
        else:
            print(bookform.errors)
        if authorform.is_valid():
            author = authorform.save()
            author.save()
            return redirect('addBooks')
        else:
            print(authorform.errors)
    book = addBookForm()   
    author =addAuthorForm()
    return render(request,'Books/addbook.html',{'book':book,'author':author})

@login_required
def viewBooks(request):
    books = Books.objects.all()
    return render(request,'Books/viewBooks.html',{'books':books})

@login_required
def deleteBook(request,id):
    try:
        b=Books.objects.get(id=id)
        b.delete() #or b.is_active = False  or b.is_deleted = 0 if data is not to be deleted
    except ObjectDoesNotExist:
        print('')
    return redirect('viewBooks')

@login_required
def updateBook(request,id):
    e = Books.objects.get(id=id)
    book = addBookForm()   
    author =addAuthorForm()
    if request.method == 'POST':
        form = addBookForm(request.POST,request.FILES, instance=e)
        if form.is_valid():
            form.save()
            return redirect('viewBooks')
        else:
            print(form.errors)
    book = addBookForm(instance=e)
    return render(request, 'Books/addbook.html',{'book':book,'author':author})